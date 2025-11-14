"""Dashboard component - displays statistics and charts for validation records."""

import csv
import io
import re
from datetime import datetime, timedelta

import pandas as pd  # type: ignore[import-untyped]
import plotly.express as px  # type: ignore[import-untyped]
import plotly.graph_objects as go  # type: ignore[import-untyped]
import streamlit as st

try:
    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore[import-untyped]
    from openpyxl.utils.dataframe import dataframe_to_rows  # type: ignore[import-untyped]
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

from staff_meal.models import OrderSource, Statistics, ValidationRecord
from staff_meal.storage import get_all_validation_records
from ui.services.alerts import Alert, detect_alerts
from ui.services.explanation import generate_dashboard_insights_sync
from ui.services.statistics import (
    calculate_statistics,
    get_statistics_by_operator,
    get_statistics_by_source,
)
from ui.utils import runner


def render_dashboard() -> None:
    """Render statistics dashboard with metrics and charts."""
    st.markdown(
        '<div style="text-align: center; font-size: 48px; font-weight: bold; margin: 20px 0;">📊 Tableau de bord</div>',
        unsafe_allow_html=True,
    )
    st.divider()

    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        start_date = st.date_input(
            "Date de début",
            value=datetime.now().date() - timedelta(days=30),
            key="dashboard_start_date",
        )
    with col2:
        end_date = st.date_input(
            "Date de fin",
            value=datetime.now().date(),
            key="dashboard_end_date",
        )
    with col3:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Actualiser", key="dashboard_refresh"):
            st.rerun()

    st.divider()

    with st.expander("🔍 Filtres avancés", expanded=False):
        col1, col2, col3 = st.columns(3)

        with col1:
            all_operators = set()
            temp_records = runner.run(
                get_all_validation_records(
                    start_date=datetime.combine(start_date, datetime.min.time()) if start_date else None,
                    end_date=datetime.combine(end_date, datetime.max.time()) if end_date else None,
                )
            )
            for record in temp_records:
                if record.operator:
                    all_operators.add(record.operator)

            selected_operators = st.multiselect(
                "Opérateur(s)",
                options=sorted(all_operators) if all_operators else [],
                key="dashboard_filter_operators",
            )

        with col2:
            selected_sources = st.multiselect(
                "Source(s)",
                options=["ubereats", "deliveroo"],
                key="dashboard_filter_sources",
            )

        with col3:
            error_types = st.multiselect(
                "Type(s) d'erreur",
                options=["Aucune erreur", "Articles manquants", "Quantités insuffisantes", "Quantités excessives", "Articles supplémentaires"],
                key="dashboard_filter_error_types",
            )

    st.divider()

    @st.cache_data(ttl=300)  # Cache for 5 minutes
    def _load_records(start_dt: datetime | None, end_dt: datetime | None) -> list[ValidationRecord]:
        """Load validation records with caching."""
        result = runner.run(get_all_validation_records(start_date=start_dt, end_date=end_dt))
        return result  # type: ignore[no-any-return]

    with st.spinner("Chargement des données..."):
        start_datetime = datetime.combine(start_date, datetime.min.time()) if start_date else None
        end_datetime = datetime.combine(end_date, datetime.max.time()) if end_date else None
        all_records = _load_records(start_datetime, end_datetime)

    if not all_records:
        st.info("Aucune donnée disponible pour la période sélectionnée.")
        return

    records = _apply_filters(
        all_records,
        operators=selected_operators if selected_operators else None,
        sources=selected_sources if selected_sources else None,
        error_types=error_types if error_types else None,
    )

    if not records:
        st.warning("Aucune donnée ne correspond aux filtres sélectionnés.")
        return

    active_filters = []
    if selected_operators:
        active_filters.append(f"Opérateurs: {', '.join(selected_operators)}")
    if selected_sources:
        active_filters.append(f"Sources: {', '.join(selected_sources)}")
    if error_types:
        active_filters.append(f"Types d'erreur: {', '.join(error_types)}")

    if active_filters:
        st.info(f"🔍 Filtres actifs: {' | '.join(active_filters)} | {len(records)}/{len(all_records)} enregistrements")

    stats = calculate_statistics(records)

    alerts = detect_alerts(stats, records)
    if alerts:
        st.markdown("#### 🚨 Alertes")
        for alert in alerts:
            if alert.severity == "critical":
                st.error(f"{alert.emoji} **{alert.title}**\n\n{alert.message}")
            elif alert.severity == "warning":
                st.warning(f"{alert.emoji} **{alert.title}**\n\n{alert.message}")
            else:
                st.info(f"{alert.emoji} **{alert.title}**\n\n{alert.message}")
        st.divider()

    st.markdown("#### 📈 Métriques principales")

    @st.cache_data(ttl=300)
    def _load_prev_stats(prev_start_dt: datetime | None, prev_end_dt: datetime | None) -> Statistics | None:
        """Load previous period statistics with caching."""
        if not prev_start_dt or not prev_end_dt:
            return None
        prev_records = runner.run(
            get_all_validation_records(start_date=prev_start_dt, end_date=prev_end_dt)
        )
        return calculate_statistics(prev_records) if prev_records else None

    period_days = (end_date - start_date).days if end_date and start_date else 30
    prev_start_date = start_date - timedelta(days=period_days) if start_date else None
    prev_end_date = start_date - timedelta(days=1) if start_date else None

    prev_stats = None
    if prev_start_date and prev_end_date:
        prev_start_datetime = datetime.combine(prev_start_date, datetime.min.time())
        prev_end_datetime = datetime.combine(prev_end_date, datetime.max.time())
        prev_stats = _load_prev_stats(prev_start_datetime, prev_end_datetime)

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        delta_orders = None
        if prev_stats:
            delta_orders = stats.total_orders - prev_stats.total_orders
        st.metric("Total commandes", stats.total_orders, delta=delta_orders)

    with col2:
        delta_complete = None
        if prev_stats:
            delta_complete = stats.complete_orders - prev_stats.complete_orders
        st.metric("Commandes complètes", stats.complete_orders, delta=delta_complete)

    with col3:
        completion_rate = (stats.complete_orders / stats.total_orders * 100) if stats.total_orders > 0 else 0.0
        delta_completion = None
        if prev_stats and prev_stats.total_orders > 0:
            prev_completion = (prev_stats.complete_orders / prev_stats.total_orders * 100)
            delta_completion = f"{completion_rate - prev_completion:+.1f}%"

        status_emoji = "🔴" if completion_rate < 90 else "🟡" if completion_rate < 95 else "🟢"
        st.metric(f"{status_emoji} Taux de complétude", f"{completion_rate:.1f}%", delta=delta_completion)

    with col4:
        delta_error = None
        if prev_stats:
            delta_error = f"{stats.error_rate - prev_stats.error_rate:+.1f}%"

        error_status_emoji = "🔴" if stats.error_rate > 20 else "🟡" if stats.error_rate > 10 else "🟢"
        st.metric(f"{error_status_emoji} Taux d'erreur", f"{stats.error_rate:.1f}%", delta=delta_error)

    st.divider()

    st.markdown("#### 💡 Recommandations IA")
    insights_key = "dashboard_ai_insights"
    if insights_key not in st.session_state:
        st.session_state[insights_key] = None

    col1, col2 = st.columns([3, 1])
    with col1:
        if st.session_state[insights_key]:
            _render_formatted_insights(st.session_state[insights_key])
        else:
            st.info("💡 Cliquez sur le bouton pour générer des recommandations basées sur vos données.")
    with col2:
        if st.button("✨ Générer", key="dashboard_generate_insights", type="primary"):
            with st.spinner("Analyse en cours..."):
                try:
                    insights = generate_dashboard_insights_sync(stats, records)
                except MissingCredentialsError:
                    st.warning(
                        "⚠️ **API Key manquante** : Veuillez configurer la clé API pour Text Generation "
                        "dans la barre latérale (section ⚙️ Celeste AI config) ou définir la variable "
                        "d'environnement pour le fournisseur."
                    )
                    insights = "Configuration de l'API requise pour générer les recommandations."
                st.session_state[insights_key] = insights
                st.rerun()

    st.divider()

    st.markdown("#### 📊 Visualisations")

    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        ["📈 Tendances", "🔍 Analyse des erreurs", "📦 Articles", "👥 Performance", "📱 Sources"]
    )

    with tab1:
        _render_trend_charts(records, stats)

    with tab2:
        _render_error_analysis_charts(records, stats)

    with tab3:
        _render_item_analysis_charts(stats)

    with tab4:
        _render_operator_performance(records)

    with tab5:
        _render_source_comparison(records)

    st.divider()
    st.markdown("#### 💾 Export des données")

    col1, col2 = st.columns(2)

    with col1:
        csv_data = _create_csv_export(records, stats)
        st.download_button(
            label="📥 Télécharger CSV",
            data=csv_data,
            file_name=f"validations_{start_date}_{end_date}.csv",
            mime="text/csv",
            key="dashboard_export_csv",
        )

    with col2:
        excel_data = _create_excel_export(records, stats)
        st.download_button(
            label="📊 Télécharger Excel",
            data=excel_data,
            file_name=f"validations_{start_date}_{end_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dashboard_export_excel",
        )


def _create_csv_export(records: list[ValidationRecord], stats: Statistics) -> str:
    """Create CSV export of validation records with calculated fields."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(
        [
            "ID",
            "Order ID",
            "Timestamp",
            "Date",
            "Heure",
            "Jour",
            "Operator",
            "Source",
            "Is Complete",
            "Nombre erreurs",
            "Expected Items",
            "Detected Items",
            "Missing Items",
            "Too Few Items",
            "Too Many Items",
            "Extra Items",
        ]
    )

    for record in records:
        expected_items_str = ", ".join([f"{item.quantity}x {item.item.value}" for item in record.expected_order.items])
        detected_items_str = ", ".join([f"{item.quantity}x {item.item.value}" for item in record.detected_order.items])
        missing_items_str = ", ".join(
            [f"{m.expected_quantity}x {m.item.value}" for m in record.comparison_result.missing_items]
        )
        too_few_str = ", ".join(
            [f"{m.expected_quantity}x {m.item.value}" for m in record.comparison_result.too_few_items]
        )
        too_many_str = ", ".join(
            [f"{m.expected_quantity}x {m.item.value}" for m in record.comparison_result.too_many_items]
        )
        extra_items_str = ", ".join(
            [f"{item.quantity}x {item.item.value}" for item in record.comparison_result.extra_items]
        )

        error_count = (
            len(record.comparison_result.missing_items)
            + len(record.comparison_result.too_few_items)
            + len(record.comparison_result.too_many_items)
            + len(record.comparison_result.extra_items)
        )

        day_names = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
        day_name = day_names[record.timestamp.weekday()]

        writer.writerow(
            [
                record.id or "",
                record.order_id,
                record.timestamp.isoformat(),
                record.timestamp.date().isoformat(),
                f"{record.timestamp.hour:02d}:{record.timestamp.minute:02d}",
                day_name,
                record.operator or "",
                record.expected_order.source.value,
                "Oui" if record.is_complete else "Non",
                error_count,
                expected_items_str,
                detected_items_str,
                missing_items_str,
                too_few_str,
                too_many_str,
                extra_items_str,
            ]
        )

    writer.writerow([])
    writer.writerow(["RÉSUMÉ"])
    writer.writerow(["Total commandes", stats.total_orders])
    writer.writerow(["Commandes complètes", stats.complete_orders])
    writer.writerow(["Taux de complétude", f"{(stats.complete_orders / stats.total_orders * 100) if stats.total_orders > 0 else 0.0:.1f}%"])
    writer.writerow(["Taux d'erreur", f"{stats.error_rate:.1f}%"])

    return output.getvalue()


def _create_excel_export(records: list[ValidationRecord], stats: Statistics) -> bytes:
    """Create Excel export of validation records."""
    if not HAS_EXCEL:
        # Fallback to CSV if openpyxl not available
        csv_data = _create_csv_export(records, stats)
        return csv_data.encode("utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "Validations"

    data_rows = []
    headers = [
        "ID",
        "Order ID",
        "Timestamp",
        "Date",
        "Heure",
        "Jour",
        "Operator",
        "Source",
        "Is Complete",
        "Nombre erreurs",
        "Expected Items",
        "Detected Items",
        "Missing Items",
        "Too Few Items",
        "Too Many Items",
        "Extra Items",
    ]
    data_rows.append(headers)

    for record in records:
        expected_items_str = ", ".join([f"{item.quantity}x {item.item.value}" for item in record.expected_order.items])
        detected_items_str = ", ".join([f"{item.quantity}x {item.item.value}" for item in record.detected_order.items])
        missing_items_str = ", ".join(
            [f"{m.expected_quantity}x {m.item.value}" for m in record.comparison_result.missing_items]
        )
        too_few_str = ", ".join(
            [f"{m.expected_quantity}x {m.item.value}" for m in record.comparison_result.too_few_items]
        )
        too_many_str = ", ".join(
            [f"{m.expected_quantity}x {m.item.value}" for m in record.comparison_result.too_many_items]
        )
        extra_items_str = ", ".join(
            [f"{item.quantity}x {item.item.value}" for item in record.comparison_result.extra_items]
        )

        error_count = (
            len(record.comparison_result.missing_items)
            + len(record.comparison_result.too_few_items)
            + len(record.comparison_result.too_many_items)
            + len(record.comparison_result.extra_items)
        )

        day_names = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
        day_name = day_names[record.timestamp.weekday()]

        data_rows.append([
            str(record.id) if record.id else "",
            record.order_id,
            record.timestamp.isoformat(),
            record.timestamp.date().isoformat(),
            f"{record.timestamp.hour:02d}:{record.timestamp.minute:02d}",
            day_name,
            record.operator or "",
            record.expected_order.source.value,
            "Oui" if record.is_complete else "Non",
            str(error_count),
            expected_items_str,
            detected_items_str,
            missing_items_str,
            too_few_str,
            too_many_str,
            extra_items_str,
        ])

    for row in data_rows:
        ws.append(row)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    summary_row = len(data_rows) + 2
    ws.cell(row=summary_row, column=1, value="RÉSUMÉ").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value="Total commandes")
    ws.cell(row=summary_row + 1, column=2, value=stats.total_orders)
    ws.cell(row=summary_row + 2, column=1, value="Commandes complètes")
    ws.cell(row=summary_row + 2, column=2, value=stats.complete_orders)
    completion_rate = (stats.complete_orders / stats.total_orders * 100) if stats.total_orders > 0 else 0.0
    ws.cell(row=summary_row + 3, column=1, value="Taux de complétude")
    ws.cell(row=summary_row + 3, column=2, value=f"{completion_rate:.1f}%")
    ws.cell(row=summary_row + 4, column=1, value="Taux d'erreur")
    ws.cell(row=summary_row + 4, column=2, value=f"{stats.error_rate:.1f}%")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _render_formatted_insights(insights_text: str) -> None:
    """Render AI insights in formatted cards with improved parsing."""
    insights_text = re.sub(r"\n{3,}", "\n\n", insights_text.strip())

    lines = insights_text.split("\n")
    recommendations = []
    current_rec = ""

    emoji_pattern = re.compile(r"^[🔴🟡🟢⚠️📌💡]")
    bullet_pattern = re.compile(r"^[-•*]\s*")
    number_pattern = re.compile(r"^\d+[\.)]\s*")
    critical_pattern = re.compile(r"^(CRITIQUE|CRITICAL|URGENT|URGENTE)", re.IGNORECASE)
    attention_pattern = re.compile(r"^(ATTENTION|WARNING|ALERTE|FOCUS)", re.IGNORECASE)
    action_pattern = re.compile(r"^(ACTION|RECOMMANDATION|SUGGESTION)", re.IGNORECASE)

    for line in lines:
        line = line.strip()
        if not line:
            if current_rec:
                recommendations.append(current_rec.strip())
                current_rec = ""
            continue

        is_new_rec = (
            emoji_pattern.match(line)
            or bullet_pattern.match(line)
            or number_pattern.match(line)
            or critical_pattern.match(line)
            or attention_pattern.match(line)
            or action_pattern.match(line)
        )

        if is_new_rec:
            if current_rec:
                recommendations.append(current_rec.strip())
            current_rec = line
        elif current_rec:
            current_rec += " " + line
        else:
            current_rec = line

    if current_rec:
        recommendations.append(current_rec.strip())

    for rec in recommendations:
        if not rec or len(rec) < 5:  # Skip very short recommendations
            continue

        rec_upper = rec.upper()

        if "🔴" in rec or "CRITIQUE" in rec_upper or "CRITICAL" in rec_upper or "URGENT" in rec_upper:
            st.error(rec)
        elif "🟡" in rec or "ATTENTION" in rec_upper or "WARNING" in rec_upper or "FOCUS" in rec_upper:
            st.warning(rec)
        elif "🟢" in rec or "OK" in rec_upper or "SUCCÈS" in rec_upper:
            st.success(rec)
        elif "💡" in rec or "ACTION" in rec_upper or "RECOMMANDATION" in rec_upper or "SUGGESTION" in rec_upper:
            st.info(rec)
        elif "⚠️" in rec or "ALERTE" in rec_upper:
            st.warning(rec)
        else:
            st.markdown(f"📌 {rec}")


def _render_trend_charts(records: list[ValidationRecord], stats: Statistics) -> None:
    """Render trend analysis charts."""
    if not records:
        st.info("Aucune donnée pour afficher les tendances.")
        return

    records_by_date: dict[str, list[ValidationRecord]] = {}
    for record in records:
        date_key = record.timestamp.date().isoformat()
        if date_key not in records_by_date:
            records_by_date[date_key] = []
        records_by_date[date_key].append(record)

    dates = []
    completion_rates = []
    error_counts = []
    for date_key in sorted(records_by_date.keys()):
        daily_records = records_by_date[date_key]
        daily_stats = calculate_statistics(daily_records)
        dates.append(date_key)
        completion_rate = (
            (daily_stats.complete_orders / daily_stats.total_orders * 100) if daily_stats.total_orders > 0 else 0.0
        )
        completion_rates.append(completion_rate)
        error_counts.append(daily_stats.total_orders - daily_stats.complete_orders)

    if dates:
        df_trend = pd.DataFrame(
            {
                "Date": pd.to_datetime(dates),
                "Taux de complétude (%)": completion_rates,
                "Nombre d'erreurs": error_counts,
            }
        )

        fig_completion = go.Figure()
        fig_completion.add_trace(
            go.Scatter(
                x=df_trend["Date"],
                y=df_trend["Taux de complétude (%)"],
                mode="lines+markers",
                fill="tozeroy",
                name="Taux de complétude",
                line=dict(color="#00cc96", width=3),
                marker=dict(size=6),
            )
        )
        fig_completion.add_hline(
            y=95, line_dash="dash", line_color="orange", annotation_text="Objectif: 95%", annotation_position="right"
        )
        fig_completion.update_layout(
            title="📈 Évolution du taux de complétude",
            xaxis_title="Date",
            yaxis_title="Taux de complétude (%)",
            height=350,
            hovermode="x unified",
        )
        st.plotly_chart(fig_completion, use_container_width=True)

        fig_errors = go.Figure()
        fig_errors.add_trace(
            go.Bar(
                x=df_trend["Date"],
                y=df_trend["Nombre d'erreurs"],
                name="Erreurs",
                marker_color="#ef553b",
            )
        )
        fig_errors.update_layout(
            title="📉 Nombre d'erreurs par jour",
            xaxis_title="Date",
            yaxis_title="Nombre d'erreurs",
            height=300,
        )
        st.plotly_chart(fig_errors, use_container_width=True)


def _render_error_analysis_charts(records: list[ValidationRecord], stats: Statistics) -> None:
    """Render error analysis charts."""
    missing_count = sum(len(r.comparison_result.missing_items) for r in records if not r.is_complete)
    too_few_count = sum(len(r.comparison_result.too_few_items) for r in records if not r.is_complete)
    too_many_count = sum(len(r.comparison_result.too_many_items) for r in records if not r.is_complete)
    extra_count = sum(len(r.comparison_result.extra_items) for r in records if not r.is_complete)

    if missing_count + too_few_count + too_many_count + extra_count > 0:
        error_types = {
            "Articles manquants": missing_count,
            "Quantités insuffisantes": too_few_count,
            "Quantités excessives": too_many_count,
            "Articles supplémentaires": extra_count,
        }
        error_types = {k: v for k, v in error_types.items() if v > 0}

        fig_pie = px.pie(
            values=list(error_types.values()),
            names=list(error_types.keys()),
            title="🥧 Répartition des types d'erreurs",
            color_discrete_sequence=px.colors.qualitative.Set3,
        )
        fig_pie.update_traces(textposition="inside", textinfo="percent+label")
        fig_pie.update_layout(height=400)
        st.plotly_chart(fig_pie, use_container_width=True)

    if stats.errors_by_hour and stats.errors_by_day:
        day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        hours = list(range(24))

        heatmap_data: dict[str, dict[int, int]] = {day: {hour: 0 for hour in hours} for day in day_order}

        for record in records:
            if not record.is_complete:
                day_name = day_order[record.timestamp.weekday()]
                hour = record.timestamp.hour
                heatmap_data[day_name][hour] = heatmap_data[day_name].get(hour, 0) + 1

        z_data = [[heatmap_data[day][hour] for hour in hours] for day in day_order]

        fig_heatmap = go.Figure(
            data=go.Heatmap(
                z=z_data,
                x=[f"{h}h" for h in hours],
                y=day_order,
                colorscale="Reds",
                showscale=True,
                text=z_data,
                texttemplate="%{text}",
                textfont={"size": 10},
            )
        )
        fig_heatmap.update_layout(
            title="🔥 Heatmap: Erreurs par jour et heure",
            xaxis_title="Heure",
            yaxis_title="Jour",
            height=400,
        )
        st.plotly_chart(fig_heatmap, use_container_width=True)

    if stats.errors_by_hour:
        hours = list(range(24))
        error_counts = [stats.errors_by_hour.get(hour, 0) for hour in hours]

        fig_hours = go.Figure()
        colors = ["#ef553b" if count > 0 else "#bdbdbd" for count in error_counts]
        fig_hours.add_trace(
            go.Bar(
                x=[f"{h}h" for h in hours],
                y=error_counts,
                marker_color=colors,
                text=error_counts,
                textposition="outside",
            )
        )
        fig_hours.update_layout(
            title="⏰ Erreurs par heure de la journée",
            xaxis_title="Heure",
            yaxis_title="Nombre d'erreurs",
            height=300,
        )
        st.plotly_chart(fig_hours, use_container_width=True)

    if stats.errors_by_day:
        day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        days = [day for day in day_order if day in stats.errors_by_day]
        error_counts = [stats.errors_by_day.get(day, 0) for day in days]

        day_names_fr = {
            "Monday": "Lundi",
            "Tuesday": "Mardi",
            "Wednesday": "Mercredi",
            "Thursday": "Jeudi",
            "Friday": "Vendredi",
            "Saturday": "Samedi",
            "Sunday": "Dimanche",
        }
        days_fr = [day_names_fr.get(day, day) for day in days]

        fig_days = go.Figure()
        fig_days.add_trace(
            go.Bar(
                x=days_fr,
                y=error_counts,
                marker_color="#ef553b",
                text=error_counts,
                textposition="outside",
            )
        )
        fig_days.update_layout(
            title="📅 Erreurs par jour de la semaine",
            xaxis_title="Jour",
            yaxis_title="Nombre d'erreurs",
            height=300,
        )
        st.plotly_chart(fig_days, use_container_width=True)


def _render_item_analysis_charts(stats: Statistics) -> None:
    """Render item analysis charts."""
    if not stats.most_forgotten_items:
        st.info("Aucun article oublié détecté.")
        return

    top_items = stats.most_forgotten_items[:10]
    items_data = [{"Article": item.value, "Nombre d'oublis": count} for item, count in top_items]
    df_items = pd.DataFrame(items_data)

    fig_items = go.Figure()
    fig_items.add_trace(
        go.Bar(
            y=df_items["Article"],
            x=df_items["Nombre d'oublis"],
            orientation="h",
            marker_color="#ef553b",
            text=df_items["Nombre d'oublis"],
            textposition="outside",
        )
    )
    fig_items.update_layout(
        title="🔴 Top 10 des articles les plus souvent oubliés",
        xaxis_title="Nombre d'oublis",
        yaxis_title="Article",
        height=400,
    )
    st.plotly_chart(fig_items, use_container_width=True)

    st.markdown("**📋 Détail des articles oubliés**")
    st.dataframe(df_items, use_container_width=True, hide_index=True)


def _render_operator_performance(records: list[ValidationRecord]) -> None:
    """Render operator performance analysis."""
    if not records:
        st.info("Aucune donnée pour afficher la performance des opérateurs.")
        return

    operator_stats = get_statistics_by_operator(records)

    if not operator_stats:
        st.info("Aucun opérateur trouvé dans les données.")
        return

    operators_data = []
    for operator, stats_op in operator_stats.items():
        completion_rate = (
            (stats_op.complete_orders / stats_op.total_orders * 100) if stats_op.total_orders > 0 else 0.0
        )
        operators_data.append({
            "Opérateur": operator,
            "Total commandes": stats_op.total_orders,
            "Commandes complètes": stats_op.complete_orders,
            "Taux de complétude (%)": completion_rate,
            "Taux d'erreur (%)": stats_op.error_rate,
        })

    df_operators = pd.DataFrame(operators_data)
    df_operators = df_operators.sort_values("Taux de complétude (%)", ascending=False)

    fig_completion = go.Figure()
    colors = [
        "#00cc96" if rate >= 95 else "#ffa15a" if rate >= 90 else "#ef553b"
        for rate in df_operators["Taux de complétude (%)"]
    ]
    fig_completion.add_trace(
        go.Bar(
            x=df_operators["Opérateur"],
            y=df_operators["Taux de complétude (%)"],
            marker_color=colors,
            text=[f"{rate:.1f}%" for rate in df_operators["Taux de complétude (%)"]],
            textposition="outside",
        )
    )
    fig_completion.add_hline(
        y=95, line_dash="dash", line_color="orange", annotation_text="Objectif: 95%", annotation_position="right"
    )
    fig_completion.update_layout(
        title="📊 Taux de complétude par opérateur",
        xaxis_title="Opérateur",
        yaxis_title="Taux de complétude (%)",
        height=400,
    )
    st.plotly_chart(fig_completion, use_container_width=True)

    fig_errors = go.Figure()
    error_colors = [
        "#ef553b" if rate > 20 else "#ffa15a" if rate > 10 else "#00cc96"
        for rate in df_operators["Taux d'erreur (%)"]
    ]
    fig_errors.add_trace(
        go.Bar(
            x=df_operators["Opérateur"],
            y=df_operators["Taux d'erreur (%)"],
            marker_color=error_colors,
            text=[f"{rate:.1f}%" for rate in df_operators["Taux d'erreur (%)"]],
            textposition="outside",
        )
    )
    fig_errors.update_layout(
        title="⚠️ Taux d'erreur par opérateur",
        xaxis_title="Opérateur",
        yaxis_title="Taux d'erreur (%)",
        height=400,
    )
    st.plotly_chart(fig_errors, use_container_width=True)

    st.markdown("**📋 Tableau détaillé par opérateur**")
    st.dataframe(df_operators, use_container_width=True, hide_index=True)

    st.markdown("**🔴 Articles les plus oubliés par opérateur**")
    for operator, stats_op in sorted(operator_stats.items(), key=lambda x: x[1].error_rate, reverse=True):
        if stats_op.most_forgotten_items:
            with st.expander(f"{operator} ({stats_op.error_rate:.1f}% erreurs)", expanded=False):
                items_data = [
                    {"Article": item.value, "Nombre d'oublis": count}
                    for item, count in stats_op.most_forgotten_items[:5]
                ]
                if items_data:
                    df_items = pd.DataFrame(items_data)
                    st.dataframe(df_items, use_container_width=True, hide_index=True)


def _render_source_comparison(records: list[ValidationRecord]) -> None:
    """Render source comparison analysis (UberEats vs Deliveroo)."""
    if not records:
        st.info("Aucune donnée pour comparer les sources.")
        return

    source_stats = get_statistics_by_source(records)

    if not source_stats:
        st.info("Aucune source trouvée dans les données.")
        return

    sources_data = []
    for source, stats_source in source_stats.items():
        completion_rate = (
            (stats_source.complete_orders / stats_source.total_orders * 100) if stats_source.total_orders > 0 else 0.0
        )
        sources_data.append({
            "Source": source.upper(),
            "Total commandes": stats_source.total_orders,
            "Commandes complètes": stats_source.complete_orders,
            "Taux de complétude (%)": completion_rate,
            "Taux d'erreur (%)": stats_source.error_rate,
        })

    df_sources = pd.DataFrame(sources_data)

    fig_comparison = go.Figure()

    fig_comparison.add_trace(
        go.Bar(
            name="Taux de complétude",
            x=df_sources["Source"],
            y=df_sources["Taux de complétude (%)"],
            marker_color="#00cc96",
            text=[f"{rate:.1f}%" for rate in df_sources["Taux de complétude (%)"]],
            textposition="outside",
        )
    )

    fig_comparison.add_trace(
        go.Bar(
            name="Taux d'erreur",
            x=df_sources["Source"],
            y=df_sources["Taux d'erreur (%)"],
            marker_color="#ef553b",
            text=[f"{rate:.1f}%" for rate in df_sources["Taux d'erreur (%)"]],
            textposition="outside",
            yaxis="y2",
        )
    )

    fig_comparison.update_layout(
        title="📊 Comparaison UberEats vs Deliveroo",
        xaxis_title="Source",
        yaxis=dict(title="Taux de complétude (%)", side="left"),
        yaxis2=dict(title="Taux d'erreur (%)", side="right", overlaying="y"),
        barmode="group",
        height=400,
    )
    st.plotly_chart(fig_comparison, use_container_width=True)

    fig_pie = px.pie(
        df_sources,
        values="Total commandes",
        names="Source",
        title="🥧 Répartition des commandes par source",
        color_discrete_sequence=px.colors.qualitative.Set2,
    )
    fig_pie.update_traces(textposition="inside", textinfo="percent+label+value")
    fig_pie.update_layout(height=400)
    st.plotly_chart(fig_pie, use_container_width=True)

    st.markdown("**📋 Tableau comparatif détaillé**")
    st.dataframe(df_sources, use_container_width=True, hide_index=True)

    st.markdown("**🔍 Répartition des types d'erreurs par source**")
    for source, stats_source in source_stats.items():
        source_records = [r for r in records if r.expected_order.source.value == source]

        missing_count = sum(len(r.comparison_result.missing_items) for r in source_records if not r.is_complete)
        too_few_count = sum(len(r.comparison_result.too_few_items) for r in source_records if not r.is_complete)
        too_many_count = sum(len(r.comparison_result.too_many_items) for r in source_records if not r.is_complete)
        extra_count = sum(len(r.comparison_result.extra_items) for r in source_records if not r.is_complete)

        if missing_count + too_few_count + too_many_count + extra_count > 0:
            error_types_data = {
                "Type d'erreur": ["Articles manquants", "Quantités insuffisantes", "Quantités excessives", "Articles supplémentaires"],
                "Nombre": [missing_count, too_few_count, too_many_count, extra_count],
            }
            df_error_types = pd.DataFrame(error_types_data)
            df_error_types = df_error_types[df_error_types["Nombre"] > 0]

            if not df_error_types.empty:
                with st.expander(f"{source.upper()} - Détail des erreurs", expanded=False):
                    fig_error_types = px.bar(
                        df_error_types,
                        x="Type d'erreur",
                        y="Nombre",
                        title=f"Types d'erreurs - {source.upper()}",
                        color="Nombre",
                        color_continuous_scale="Reds",
                    )
                    fig_error_types.update_layout(height=300)
                    st.plotly_chart(fig_error_types, use_container_width=True)
                    st.dataframe(df_error_types, use_container_width=True, hide_index=True)


def _apply_filters(
    records: list[ValidationRecord],
    operators: list[str] | None = None,
    sources: list[str] | None = None,
    error_types: list[str] | None = None,
) -> list[ValidationRecord]:
    """Apply filters to validation records.

    Args:
        records: List of validation records to filter.
        operators: Optional list of operator names to filter by.
        sources: Optional list of source names to filter by (ubereats, deliveroo).
        error_types: Optional list of error types to filter by.

    Returns:
        Filtered list of validation records.
    """
    filtered = records

    if operators:
        filtered = [r for r in filtered if r.operator in operators]

    if sources:
        source_enums = []
        if "ubereats" in sources:
            source_enums.append(OrderSource.UBER_EATS)
        if "deliveroo" in sources:
            source_enums.append(OrderSource.DELIVEROO)
        if source_enums:
            filtered = [r for r in filtered if r.expected_order.source in source_enums]

    if error_types:
        filtered_by_error = []
        for record in filtered:
            should_include = False

            if "Aucune erreur" in error_types and record.is_complete:
                should_include = True
            elif not record.is_complete:
                if "Articles manquants" in error_types and record.comparison_result.missing_items:
                    should_include = True
                if "Quantités insuffisantes" in error_types and record.comparison_result.too_few_items:
                    should_include = True
                if "Quantités excessives" in error_types and record.comparison_result.too_many_items:
                    should_include = True
                if "Articles supplémentaires" in error_types and record.comparison_result.extra_items:
                    should_include = True

            if should_include:
                filtered_by_error.append(record)
        filtered = filtered_by_error

    return filtered


__all__ = ["render_dashboard"]
